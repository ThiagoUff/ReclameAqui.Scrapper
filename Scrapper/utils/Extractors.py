import json
from openpyxl import Workbook


def gerar_arquivo_excel(documents, filename):
    try:
        # Criação do objeto Workbook
        wb = Workbook()

        # Seleciona a planilha ativa
        ws = wb.active
        ws['A1'] = 'id'
        ws['B1'] = 'titulo'
        ws['C1'] = 'status'
        ws['D1'] = 'data_criacao'
        ws['E1'] = 'reclamacao'
        ws['F1'] = 'estado'
        ws['G1'] = 'cidade'
        ws['H1'] = 'categoria'
        ws['I1'] = 'problema'
        ws['J1'] = 'produto'
        ws['K1'] = 'voltaria_fazer_negocio'
        ws['L1'] = 'nota'
        ws['M1'] = 'respondido_em'
        ws['N1'] = 'sla'
        index = 2
        for document in documents:
            anwser = None
            sla= None
            if document['interaction_items']:
                anwser = list(filter(
                    lambda x: x['interaction_type'] == "ANSWER", document['interaction_items']))[0]['interaction_date']
                result = anwser- document['data_criacao']
                
                # Gerar a saída formatada
                sla = result.total_seconds()

            ws[f'A{index}'] = document['id']
            ws[f'B{index}'] = document['titulo']
            ws[f'C{index}'] = document['status']
            ws[f'D{index}'] = document['data_criacao']
            ws[f'E{index}'] = document['reclamacao']
            ws[f'F{index}'] = document['estado']
            ws[f'G{index}'] = document['cidade']
            ws[f'H{index}'] = document['categoria']
            ws[f'I{index}'] = document['problema']
            ws[f'J{index}'] = document['produto']
            ws[f'K{index}'] = document['voltaria_fazer_negocio']
            ws[f'L{index}'] = document['nota']
            ws[f'M{index}'] = anwser
            ws[f'N{index}'] = sla

            index += 1

        wb.save(f'scrapper-dataset/excel/{filename}.xlsx')
    except Exception as excecao:
        print(excecao)


def gerar_arquivo_json(documents, filename):
    try:
        json_documents = []
        for document in documents:
            json_document = json.dumps(document, default=str)
            json_documents.append(json_document)
        with open(f'scrapper-dataset/json/{filename}.json', 'w') as arquivo:
            json.dump(json_documents, arquivo)
    except Exception as excecao:
        print(excecao)
